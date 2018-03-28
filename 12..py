# -*- coding:utf-8 -*-
import openpyxl  
  
wb = openpyxl.load_workbook('d:\\bg.xlsx')  
  
#��ȡworkbook�����еı��  
sheets = wb.get_sheet_names()  
  
print(sheets)  
  
#ѭ����������sheet  
for i in range(len(sheets)):  
    sheet= wb.get_sheet_by_name(sheets[i])  
      
    print('\n\n��'+str(i+1)+'��sheet: ' + sheet.title+'->>>')  
  
    for r in range(1,sheet.max_row+1):  
        if r == 1:  
            print('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))  
        else:  
            print(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) for c in range(1,sheet.max_column+1)] )) 